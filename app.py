import base64
import io
import os
import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

# ==========================================
# 1. PAGE SETUP & ASSET LOADER
# ==========================================
st.set_page_config(
    page_title="PDFtoXL",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

def get_base64_image(image_base_name):
    """Detects either .png or .jpg and returns the base64 string with correct MIME type."""
    for ext, mime in [(".png", "image/png"), (".jpg", "image/jpeg"), (".jpeg", "image/jpeg")]:
        full_path = image_base_name + ext
        if os.path.exists(full_path):
            with open(full_path, "rb") as img_file:
                encoded = base64.b64encode(img_file.read()).decode()
                return f"data:{mime};base64,{encoded}"
    return ""

bg_data_uri = get_base64_image("background")
logo_data_uri = get_base64_image("logo")

# ==========================================
# 2. MOBILE-FRIENDLY RESPONSIVE UI
# ==========================================
st.markdown(
    f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;800;900&display=swap');

    /* Global Fullscreen Background */
    .stApp {{
        background: {f'linear-gradient(rgba(10, 17, 30, 0.45), rgba(10, 17, 30, 0.55)), url("{bg_data_uri}")' if bg_data_uri else '#0f172a'};
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        font-family: 'Poppins', sans-serif;
        color: #FFFFFF;
    }}

    header, footer {{visibility: hidden;}}
    #MainMenu {{visibility: hidden;}}

    /* Responsive Glass Card */
    .block-container {{
        max-width: 740px !important;
        margin: 2vh auto 2rem auto !important;
        padding: 2.5rem 1.5rem !important;
        background: rgba(15, 23, 42, 0.32) !important;
        backdrop-filter: blur(4px) !important;
        -webkit-backdrop-filter: blur(4px) !important;
        border: 1px solid rgba(255, 255, 255, 0.22) !important;
        border-radius: 24px !important;
        box-shadow: 0 20px 50px rgba(0, 0, 0, 0.45) !important;
    }}

    /* Responsive Brand Header */
    .brand-header {{
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        margin-bottom: 1.5rem;
        text-align: center;
    }}
    .brand-logo {{
        width: 75px;
        height: auto;
        margin-bottom: 10px;
        filter: drop-shadow(0 6px 14px rgba(0,0,0,0.5));
    }}
    
    .brand-title {{
        font-size: clamp(3rem, 12vw, 5.5rem) !important;
        font-weight: 900 !important;
        line-height: 1;
        letter-spacing: -2px;
        margin: 0;
        background: linear-gradient(135deg, #FFFFFF 30%, #38BDF8 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        filter: drop-shadow(0 4px 12px rgba(0,0,0,0.4));
    }}
    .brand-subtitle {{
        font-size: clamp(0.8rem, 3vw, 0.95rem);
        color: #E2E8F0;
        margin-top: 8px;
        letter-spacing: 0.5px;
    }}

    /* Responsive File Uploader */
    [data-testid="stFileUploader"] {{
        width: 100%;
        margin-top: 10px;
    }}
    [data-testid="stFileUploader"] section {{
        background: rgba(255, 255, 255, 0.05) !important;
        border: 2px dashed rgba(56, 189, 248, 0.55) !important;
        border-radius: 18px !important;
        padding: 2rem 1rem !important;
        cursor: pointer;
    }}

    /* Action Buttons */
    .stButton > button {{
        background: linear-gradient(135deg, #0284C7 0%, #0EA5E9 100%) !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        font-size: 1.05rem !important;
        border: none !important;
        border-radius: 14px !important;
        padding: 0.85rem 1.5rem !important;
        box-shadow: 0 10px 20px rgba(14, 165, 233, 0.35) !important;
        margin-top: 10px;
    }}

    /* Download Box Button */
    [data-testid="stDownloadButton"] > button {{
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        font-size: 1.05rem !important;
        border: none !important;
        border-radius: 14px !important;
        padding: 0.9rem 1.5rem !important;
        box-shadow: 0 12px 24px rgba(16, 185, 129, 0.35) !important;
    }}

    .brand-footnote {{
        text-align: center;
        font-size: 0.78rem;
        color: #94A3B8;
        margin-top: 2rem;
        letter-spacing: 1.2px;
        text-transform: uppercase;
        font-weight: 600;
    }}
    </style>
    """,
    unsafe_allow_html=True
)

# Desktop-only JS (bypassed on mobile to prevent touch tap hijack)
components.html(
    """
    <script>
    const isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent);
    if (!isMobile) {
        const doc = window.parent.document;
        ['dragenter', 'dragover'].forEach(eventName => {
            doc.addEventListener(eventName, (e) => {
                e.preventDefault();
                e.stopPropagation();
            }, false);
        });

        ['dragleave', 'drop'].forEach(eventName => {
            doc.addEventListener(eventName, (e) => {
                e.preventDefault();
                e.stopPropagation();
                if (eventName === 'drop' && e.dataTransfer.files.length > 0) {
                    const input = doc.querySelector('input[type="file"]');
                    if (input) {
                        input.files = e.dataTransfer.files;
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                }
            }, false);
        });

        doc.addEventListener('paste', (e) => {
            const items = (e.clipboardData || e.originalEvent.clipboardData).items;
            for (let item of items) {
                if (item.kind === 'file') {
                    const file = item.getAsFile();
                    if (file && file.type === 'application/pdf') {
                        const dataTransfer = new DataTransfer();
                        dataTransfer.items.add(file);
                        const input = doc.querySelector('input[type="file"]');
                        if (input) {
                            input.files = dataTransfer.files;
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                        }
                    }
                }
            }
        });
    }
    </script>
    """,
    height=0,
    width=0
)

# Branding Section
logo_html = f'<img src="{logo_data_uri}" class="brand-logo"/>' if logo_data_uri else ''
st.markdown(
    f"""
    <div class="brand-header">
        {logo_html}
        <h1 class="brand-title">PDFtoXL</h1>
        <p class="brand-subtitle">PDF to Excel assistance (Prototype Version)</p>
    </div>
    """,
    unsafe_allow_html=True
)

# ==========================================
# 3. CORE PROCESSING ENGINE
# ==========================================
def get_strict_vertical_boundaries(page):
    """Captures black, light grey, thin stroke, rect and curve vertical boundaries."""
    v_lines = []
    
    # 1. Capture all drawn lines & edges (including grey/colored edges)
    all_edges = getattr(page, 'edges', []) or page.lines
    for edge in all_edges:
        if abs(edge['x0'] - edge['x1']) <= 3.0 and abs(edge['bottom'] - edge['top']) >= 8:
            v_lines.append((edge['x0'] + edge['x1']) / 2.0)

    # 2. Capture vertical strokes rendered as rectangles
    for rect in page.rects:
        if rect.get('height', 0) >= 8:
            if rect.get('width', 0) <= 5:
                v_lines.append((rect['x0'] + rect['x1']) / 2.0)
            else:
                v_lines.append(rect['x0'])
                v_lines.append(rect['x1'])

    # 3. Capture vector curve vertical strokes
    for curve in getattr(page, 'curves', []):
        if abs(curve['x0'] - curve['x1']) <= 3.0 and abs(curve['bottom'] - curve['top']) >= 8:
            v_lines.append((curve['x0'] + curve['x1']) / 2.0)

    # Deduplicate lines within 4px tolerance
    v_lines = sorted(v_lines)
    unique_lines = []
    for x in v_lines:
        if not unique_lines or abs(x - unique_lines[-1]) > 4.0:
            unique_lines.append(x)
            
    if len(unique_lines) >= 2:
        return [(unique_lines[i], unique_lines[i+1]) for i in range(len(unique_lines)-1)]
    return None

def extract_strict_x_grid(page):
    words = page.extract_words(
        x_tolerance=2,
        y_tolerance=2,
        keep_blank_chars=False,
        use_text_flow=False
    )
    
    if not words:
        return None

    # Exact Y-axis line grouping preserved
    words = sorted(words, key=lambda w: (w['top'], w['x0']))
    lines = []
    curr_line = [words[0]]

    for w in words[1:]:
        if abs(w['top'] - curr_line[-1]['top']) <= 3.5:
            curr_line.append(w)
        else:
            lines.append(sorted(curr_line, key=lambda x: x['x0']))
            curr_line = [w]
    if curr_line:
        lines.append(sorted(curr_line, key=lambda x: x['x0']))

    col_intervals = get_strict_vertical_boundaries(page)

    if not col_intervals:
        x_positions = sorted([w['x0'] for w in words])
        col_clusters = []
        for x in x_positions:
            if not col_clusters:
                col_clusters.append([x])
            else:
                if abs(x - (sum(col_clusters[-1]) / len(col_clusters[-1]))) < 16.0:
                    col_clusters[-1].append(x)
                else:
                    col_clusters.append([x])
        col_mids = [sum(cluster) / len(cluster) for cluster in col_clusters]
    else:
        col_mids = None

    num_cols = len(col_intervals) if col_intervals else len(col_mids)
    grid = []

    for line in lines:
        row = [""] * num_cols
        for w in line:
            w_mid_x = (w['x0'] + w['x1']) / 2.0
            
            if col_intervals:
                col_idx = None
                for idx, (x_left, x_right) in enumerate(col_intervals):
                    if (x_left - 3.0) <= w_mid_x < (x_right + 3.0):
                        col_idx = idx
                        break
                
                if col_idx is None:
                    col_idx = min(range(num_cols), key=lambda i: min(abs(w_mid_x - col_intervals[i][0]), abs(w_mid_x - col_intervals[i][1])))
            else:
                col_idx = min(range(num_cols), key=lambda i: abs(w['x0'] - col_mids[i]))

            if row[col_idx]:
                row[col_idx] += " " + w['text']
            else:
                row[col_idx] = w['text']

        if any(cell.strip() for cell in row):
            grid.append(row)

    if not grid:
        return None

    df = pd.DataFrame(grid)
    df = df.loc[:, (df != "").any(axis=0)]
    df.columns = [f"Col_{i+1}" for i in range(df.shape[1])]
    return df

# ==========================================
# 4. ONE-CLICK USER INTERACTION
# ==========================================
uploaded_file = st.file_uploader(
    "Choose a PDF file",
    type=["pdf"],
    label_visibility="visible"
)

if uploaded_file is not None:
    if st.button("⚡ Convert to Excel", type="primary", use_container_width=True):
        with st.spinner("Extracting strictly locked grid..."):
            try:
                pages_data = {}

                with pdfplumber.open(uploaded_file) as pdf:
                    for page_idx, page in enumerate(pdf.pages):
                        page_num = page_idx + 1
                        df_page = extract_strict_x_grid(page)
                        if df_page is not None and not df_page.empty:
                            pages_data[f"Page_{page_num}"] = df_page

                if not pages_data:
                    st.error("No table data found in this PDF.")
                else:
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        for sheet_name, df in pages_data.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                            ws = writer.sheets[sheet_name]

                            border_thin = Border(
                                left=Side(style='thin', color='D9D9D9'),
                                right=Side(style='thin', color='D9D9D9'),
                                top=Side(style='thin', color='D9D9D9'),
                                bottom=Side(style='thin', color='D9D9D9')
                            )

                            for row_idx, row in enumerate(ws.iter_rows()):
                                for cell in row:
                                    cell.border = border_thin
                                    val_str = str(cell.value or '').strip()
                                    
                                    if row_idx == 0:
                                        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                                        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    else:
                                        clean_num = val_str.replace('.', '', 1).replace(',', '', 1).replace('-', '', 1).replace('₹', '').replace('$', '').strip()
                                        if clean_num.isdigit() and len(clean_num) > 0:
                                            cell.alignment = Alignment(horizontal="right", vertical="center")
                                        else:
                                            cell.alignment = Alignment(horizontal="left", vertical="center")

                            for col in ws.columns:
                                max_len = 0
                                col_letter = get_column_letter(col[0].column)
                                for cell in col:
                                    val_str = str(cell.value or '').strip()
                                    if len(val_str) > max_len:
                                        max_len = len(val_str)
                                ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

                    excel_data = output.getvalue()
                    st.success(f"Extracted {sum(len(df) for df in pages_data.values())} rows across {len(pages_data)} page(s)!")

                    st.download_button(
                        label="📥 Download .xlsx File",
                        data=excel_data,
                        file_name=f"{uploaded_file.name.rsplit('.', 1)[0]}_converted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                    st.write("---")
                    preview_tabs = st.tabs(list(pages_data.keys()))
                    for tab, (sheet_name, df) in zip(preview_tabs, pages_data.items()):
                        with tab:
                            st.dataframe(df, use_container_width=True, hide_index=True)

            except Exception as e:
                st.error(f"Error: {e}")

st.markdown(
    '<div class="brand-footnote">Built by DHRUV, with hate.</div>',
    unsafe_allow_html=True
)
